"""Excel export: the payment-confirmation workbook an admin hands off to
decide who gets paid. Two sheets: one row per boleta with a clear
Pagar (Sí/No) column, and a fletero totals sheet -- both driven by
`status == "auto_processed"` (the pipeline's own confirmation gate), so a
flagged/unverified boleta never silently counts as payable.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import Boleta, BoletaRecord
from app.reporting.summary import build_batch_summary

BOLETAS_COLUMNS = [
    "Folio",
    "Fecha",
    "Origen",
    "Destino",
    "Fletero",
    "Tipo de viaje",
    "Tarifa",
    "Estado",
    "Pagar",
    "Excepciones",
]

HEADER_FILL = PatternFill(start_color="F2EDE4", end_color="F2EDE4", fill_type="solid")
HEADER_FONT = Font(bold=True)


def _style_header(ws, columns: list[str]) -> None:
    for col_idx, title in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")
    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18


def build_batch_xlsx(db: Session, batch_id: int) -> bytes:
    records = (
        db.query(BoletaRecord)
        .join(Boleta, BoletaRecord.boleta_id == Boleta.id)
        .filter(Boleta.batch_id == batch_id)
        .filter(BoletaRecord.reconciled_with_record_id.is_(None))  # see app/reporting/summary.py
        .order_by(BoletaRecord.id)
        .all()
    )
    summary = build_batch_summary(db, batch_id)

    wb = Workbook()

    boletas_ws = wb.active
    boletas_ws.title = "Boletas"
    _style_header(boletas_ws, BOLETAS_COLUMNS)
    for row_idx, r in enumerate(records, start=2):
        pagar = "Sí" if r.status == "auto_processed" else "No"
        boletas_ws.append(
            [
                r.folio,
                r.date,
                r.origin,
                r.destination,
                r.fletero,
                r.trip_type,
                r.tariff_amount,
                r.status,
                pagar,
                "; ".join(r.exceptions or []),
            ]
        )
        if pagar == "No":
            for col_idx in range(1, len(BOLETAS_COLUMNS) + 1):
                boletas_ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color="FDF0E0", end_color="FDF0E0", fill_type="solid"
                )
    boletas_ws.freeze_panes = "A2"

    fletero_ws = wb.create_sheet("Pago por Fletero")
    _style_header(fletero_ws, ["Fletero", "Total a pagar"])
    for fletero, amount in sorted(summary.total_payment_by_fletero.items()):
        fletero_ws.append([fletero, amount])
    fletero_ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
