from __future__ import annotations

import io

from openpyxl import load_workbook

from app.models import Batch, Boleta, BoletaRecord
from app.exports.xlsx_export import build_batch_xlsx


def _make_record(db_session, batch, *, status, folio, fletero, tariff_amount, exceptions=None):
    boleta = Boleta(
        batch_id=batch.id,
        original_filename="x.png",
        stored_path="/tmp/x.png",
        mime_type="image/png",
        page_number=1,
        sha256_hash=f"hash-{folio}",
    )
    db_session.add(boleta)
    db_session.flush()
    record = BoletaRecord(
        boleta_id=boleta.id,
        status=status,
        folio=folio,
        fletero=fletero,
        tariff_amount=tariff_amount,
        origin="Mina San Jose",
        destination="Planta Norte",
        exceptions=exceptions or [],
        field_confidences={},
    )
    db_session.add(record)
    db_session.flush()
    return record


def test_xlsx_marks_auto_processed_as_payable_and_flagged_as_not(db_session):
    batch = Batch(label="test")
    db_session.add(batch)
    db_session.flush()

    _make_record(db_session, batch, status="auto_processed", folio="B-1", fletero="Juan Perez", tariff_amount=850.0)
    _make_record(
        db_session, batch, status="needs_review", folio="B-2", fletero="Maria Lopez",
        tariff_amount=300.0, exceptions=["unknown_folio"],
    )

    xlsx_bytes = build_batch_xlsx(db_session, batch.id)
    wb = load_workbook(io.BytesIO(xlsx_bytes))

    assert wb.sheetnames == ["Boletas", "Pago por Fletero"]

    boletas_ws = wb["Boletas"]
    rows = list(boletas_ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2

    by_folio = {row[0]: row for row in rows}
    # columns: Folio, Fecha, Origen, Destino, Fletero, Tipo, Tarifa, Estado, Pagar, Excepciones
    assert by_folio["B-1"][8] == "Sí"
    assert by_folio["B-2"][8] == "No"
    assert by_folio["B-2"][9] == "unknown_folio"

    fletero_ws = wb["Pago por Fletero"]
    fletero_rows = list(fletero_ws.iter_rows(min_row=2, values_only=True))
    # Only the payable (auto_processed) boleta's fletero should appear.
    assert fletero_rows == [("Juan Perez", 850.0)]
